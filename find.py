import os
import time
from openpyxl import Workbook, load_workbook

keyList=[]
keyWithTime={}

#关键字获取函数，关键字保存到列表keyList中
def key_read(filePath):
    file=open(filePath, encoding='utf-8')
    i=1
    line= file.readline()
    while line:
        keyList.append(line.replace('\n', ''))
        line=file.readline()
        i=i+1
    file.close()
#文件保存
def save_file():
    newWb = Workbook()
    sheet = newWb.active
    number = 0
    rows = 1
    for key in keyWithTime:
        listnumber = len(keyWithTime[key])
        temp = number+listnumber
        #print(f'key:{key} and numer:{listnumber}')
        number = temp
        if listnumber > 1:
            for i in range(listnumber):
                cell_1 = sheet.cell(row=rows, column=1)
                cell_1.value = key               
                cell_2 = sheet.cell(row=rows, column=2)
                cell_2.value = keyWithTime[key][i]
                rows+= 1
        else:
            cell_1 = sheet.cell(row=rows, column=1)
            cell_1.value = key
            cell_2 = sheet.cell(row=rows, column=2)
            cell_2.value = keyWithTime[key][0]
            rows+= 1
        
        
    #print(f'number:{number}')
    newWb.save('result.xlsx')
#文件内容对比工作
def work(filePath='log.txt'):
    time=[]
    temp=[]
    file=open(filePath, encoding='utf-8')
    i=1
    line= file.readline()
    for key in keyList:
        keyWithTime[key]=[]
    while line:
        for key in keyList:
            if line.find(key)!= -1:
                time=line.split()
                time1=time[1].split(':')
                keyWithTime[key].append(format(float(time1[2]), ".3f"))
        line=file.readline()
        i=i+1
    file.close() 
    
key_read('key.txt') 
work()
save_file()
print(f'keyList: {keyList}')
print(f'keyWithTime: {keyWithTime}')
os.system('pause')
os.system('pause')