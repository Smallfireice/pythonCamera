import os
import time
from openpyxl import Workbook, load_workbook


def xslxfile(filePath):
    newWb = Workbook()
    sheet = newWb.active
    file=open(filePath, encoding='utf-8')
    line= file.readline()
    i = 1
    rows = 1
    flag = 0
    while line:
        if i % 2 == 0:
            cell_1 = sheet.cell(row=rows, column=2)
            cell_1.value = line
            print(f'{line}-----({rows, 2})')
        else:
            cell_2 = sheet.cell(row=rows, column=1)
            cell_2.value = line
            print(f'{line}-----({rows, 1})')
        line=file.readline()
        if flag != 2:
            if i == 1:
                pass
            else:
                rows+=1
        elif flag == 2:
            flag = 0
        i+=1
        flag+=1
    file.close()
    newWb.save('result.xlsx')

xslxfile("result.txt")
time.sleep(3)
print("经过3秒后自动退出")
